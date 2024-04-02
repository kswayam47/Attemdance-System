import tkinter as tk
from tkinter import simpledialog, messagebox
from PIL import Image, ImageTk
import cv2
import os
import pandas as pd
import datetime
import firebase_admin
import openpyxl
from firebase_admin import credentials, db
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
import time

# Initialize Firebase app
credentials = credentials.Certificate("service.json")
firebase_admin.initialize_app(credentials, {'databaseURL': "https://attendance-system-5b347-default-rtdb.firebaseio.com/"})

# Get a reference to the database service
root = db.reference('attendance')

# Load the trained model
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("face_trained.yml")

# Names related to ids
names = [name for name in os.listdir(r"C:\\Users\\swaya\\images")]

# Initialize and start real-time video capture
cam = cv2.VideoCapture(0)
cam.set(3, 640)  # Set video width
cam.set(4, 480)  # Set video height

# Define min window size to be recognized as a face
minW = 0.1 * cam.get(3)
minH = 0.1 * cam.get(4)

# Initialize a set to store recognized ids
recognized_ids = set()

# Function to start the attendance process
def start_attendance():
    subject = subject_entry.get()
    if not subject:
        messagebox.showerror("Error", "Please enter a subject name!")
    else:
        date_entry = simpledialog.askstring("Date Entry", "Enter the date (YYYY-MM-DD):")
        if not date_entry:
            messagebox.showerror("Error", "Please enter a valid date!")
            return
        
        messagebox.showinfo("Welcome", f"Starting attendance for {subject} on {date_entry}")

        while True:
            ret, img = cam.read()
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            faces = cv2.CascadeClassifier("C:\\OpenCV\\haaar_face.xml").detectMultiScale(
                gray,
                scaleFactor=1.2,
                minNeighbors=5,
                minSize=(int(minW), int(minH)),
            )

            for (x, y, w, h) in faces:
                cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)

                id, confidence = recognizer.predict(gray[y:y + h, x:x + w])

                if confidence < 100:
                    name = names[id]
                    confidence = " {0}%".format(round(100 - confidence))

                    if name not in recognized_ids:
                        response = messagebox.askyesno("Mark Attendance", f"Do you want to mark attendance for {name}?")
                        if response:
                            data = {"id": name, "time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
                            db.reference('attendance').push().set(data)
                            recognized_ids.add(name)
                            print(f"Attendance marked for {name} at {datetime.datetime.now()}")

                            # Write to Excel file
                            file_path = f'C:\\Users\\swaya\\.vscode\\{subject}_{date_entry}.xlsx'
                            attendance_df = pd.DataFrame({'ID': [name], 'Time': [datetime.datetime.now()]})

                            if os.path.exists(file_path):
                                book = load_workbook(file_path)
                                sheet = book.active
                            else:
                                book = openpyxl.Workbook()
                                sheet = book.active
                                sheet.append(['ID', 'Time'])

                            # Formatting the Excel file
                            sheet['A1'] = 'Subject: ' + subject
                            sheet['A2'] = 'Date: ' + date_entry
                            sheet['A1'].font = Font(bold=True)
                            sheet['A1'].alignment = Alignment(horizontal='center')

                            for cell in sheet['A2:C2'][0]:
                                cell.alignment = Alignment(horizontal='center')

                            for row in dataframe_to_rows(attendance_df, index=False, header=False):
                                sheet.append(row)

                            book.save(file_path)

                    cv2.putText(img, name, (x + 5, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)
                    cv2.putText(img, str(confidence), (x + 5, y + h - 5), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 0), 1)

            cv2.imshow('camera', img)

            k = cv2.waitKey(10)
            if k == 27:
                print("Thank you for using the attendance system!")
                cam.release()
                cv2.destroyAllWindows()
                root.destroy()  # Close the tkinter window
                break  # Exit the loop            

# UI Design
root = tk.Tk()
root.title("Attendance Marking System")

# Set window size and position it at the center of the screen
window_width = 800
window_height = 600
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x_cordinate = int((screen_width / 2) - (window_width / 2))
y_cordinate = int((screen_height / 2) - (window_height / 2))
root.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))

# Add background image to the UI
bg_image = Image.open("demo.jpg")
bg_image = bg_image.resize((window_width, window_height))  # Resize image to fit window
bg_photo = ImageTk.PhotoImage(bg_image)
bg_label = tk.Label(root, image=bg_photo)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

title_label = tk.Label(root, text="Attendance Marking System", font=("Helvetica", 24, "bold"))
title_label.pack(pady=20)

subject_frame = tk.Frame(root)
subject_label = tk.Label(subject_frame, text="Enter the subject name:")
subject_label.pack(side=tk.LEFT, padx=10)
subject_entry = tk.Entry(subject_frame, width=30, font=("Helvetica", 12))
subject_entry.pack(side=tk.LEFT, padx=10)
subject_frame.pack(pady=20)

get_started_button = tk.Button(root, text="Start Attendance", font=("Helvetica", 16, "bold"), bg="blue", fg="white", command=start_attendance)
get_started_button.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
label = tk.Label(root, text="Group ID:2C82024", font=("Helvetica", 10, "bold"), fg="white")
label.config(bg="black")
label.pack(side=tk.BOTTOM)
disclaimer_label = tk.Label(root, text="Disclaimer: Ensure proper authorization before marking attendance.", font=("Helvetica", 10, "bold"), fg="white")
disclaimer_label.config(bg="black")
disclaimer_label.pack(side=tk.BOTTOM)

root.mainloop()
