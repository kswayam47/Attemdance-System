import os
import cv2
import pandas as pd
import datetime
import firebase_admin
import openpyxl
from firebase_admin import credentials, db
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

# Initialize Firebase app
credentials = credentials.Certificate("service.json")
firebase_admin.initialize_app(credentials, {'databaseURL': "https://attendance-system-5b347-default-rtdb.firebaseio.com/"})

# Get a reference to the database service
root = db.reference('attendance')

# Load the trained model
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("face_trained.yml")

# Names related to ids
names = [name for name in os.listdir(r"C:\\Users\\swaya\\images")]

# Initialize and start real-time video capture
cam = cv2.VideoCapture(0)
cam.set(3, 640)  # Set video width
cam.set(4, 480)  # Set video height

# Define min window size to be recognized as a face
minW = 0.1 * cam.get(3)
minH = 0.1 * cam.get(4)

# Initialize a set to store recognized ids
recognized_ids = set()

subject = input("Enter the subject name: ")

while True:
    ret, img = cam.read()
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    faces = cv2.CascadeClassifier("C:\\OpenCV\\haaar_face.xml").detectMultiScale(
        gray,
        scaleFactor=1.2,
        minNeighbors=5,
        minSize=(int(minW), int(minH)),
    )

    for (x, y, w, h) in faces:
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)

        id, confidence = recognizer.predict(gray[y:y + h, x:x + w])

        if confidence < 100:
            name = names[id]
            confidence = " {0}%".format(round(100 - confidence))

            if name not in recognized_ids:
                action = input(f"Do you want to mark attendance for {name}? (Y/N): ")
                if action.upper() == 'Y':
                    data = {"id": name, "time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
                    root.push(data)
                    recognized_ids.add(name)
                    print(f"Attendance marked for {name} at {datetime.datetime.now()}")

                    # Write to Excel file
                    file_path = f'C:\\Users\\swaya\\.vscode\\{subject}.xlsx'
                    attendance_df = pd.DataFrame({'ID': [name], 'Time': [datetime.datetime.now()]})

                    if os.path.exists(file_path):
                        book = load_workbook(file_path)
                        sheet = book.active
                    else:
                        book = openpyxl.Workbook()
                        sheet = book.active
                        sheet.append(['ID', 'Time'])

                    # Formatting the Excel file
                    sheet['A1'] = 'Subject: ' + subject
                    sheet['A1'].font = Font(bold=True)
                    sheet['A1'].alignment = Alignment(horizontal='center')

                    for cell in sheet['A2:C2'][0]:
                        cell.alignment = Alignment(horizontal='center')
                    
                    for row in dataframe_to_rows(attendance_df, index=False, header=False):
                        sheet.append(row)

                    book.save(file_path)

            cv2.putText(img, name, (x + 5, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)
            cv2.putText(img, str(confidence), (x + 5, y + h - 5), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 0), 1)

    cv2.imshow('camera', img)

    k = cv2.waitKey(10)
    if k == 27:
        print("Thank you for using the attendance system!")
        break

book.save(f'C:\\Users\\swaya\\.vscode\\{subject}.xlsx')
